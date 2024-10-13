from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib
from openpyxl import load_workbook

# Opens the victim list workbook
workbook = load_workbook(filename="emailScamProject/email_list.xlsx")
sheet = workbook.active


print(sheet.cell(row=1, column=1).value)
print(sheet.cell(row=2, column=1).value)

# Example showcasing how we can iterrate through each spot and use the data from the list
for value in sheet.iter_rows(min_row=1, max_row=2, min_col=1,max_col=9,values_only=True):
    print(value)

# Notes: For testing purposes, change your desired email to your class email as I cant 2fa everyone into the scamproject email, but it should still
#   work with the app password to send emails
# Define email details
FROM = 'ascamprojectforclass@gmail.com'
TO = ["ascamprojectforclass@gmail.com"]  # lList of emails
SUBJECT = "sorryforspam"

targetName = "Victum"

with open("emailScamProject/script.html", 'r') as file:
    scriptText = file.read().format(name=targetName)

message = MIMEMultipart()
message['From'] = FROM
message['To'] = ", ".join(TO)
message['Subject'] = SUBJECT

# Attach the formatted HTML content to the email
message.attach(MIMEText(scriptText, 'html'))

# Send the mail using Gmail's SMTP server
try:
    server = smtplib.SMTP('smtp.gmail.com', 587)  # Gmail's SMTP server
    server.starttls()  # Secure the connection
    server.login(FROM, "nlrf sncc gskx lbqf")  # App Password Email
    server.sendmail(FROM, TO, message.as_string())
    print("Email sent successfully!")
except Exception as e:
    print(f"Failed to send email: {e}")
finally:
    server.quit()
